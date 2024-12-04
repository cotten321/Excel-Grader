import openpyxl
import traceback
from openpyxl.utils import get_column_letter

#prev
def grade_challenge_1_1(student_path):
    try:
        # Load workbooks and select active sheets
        student_wb = openpyxl.load_workbook(student_path)
        student_ws = student_wb.active

        # Initialize scoring variables
        score = 0
        total_points = 10
        feedback = []

        # Define expected headers and row count
        expected_headers = ["CustomerID", "FirstName", "LastName", "Email"]
        expected_row_count = 5  # 5 rows of data (not including the header row)

        # Check headers in the first row (A1:D1)
        student_headers = [student_ws.cell(row=1, column=col).value for col in range(1, 5)]
        if student_headers == expected_headers:
            score += 2  # Award 2 points for having the correct headers
        else:
            feedback.append("Headers are incorrect.")

        # Check for 5 rows of data below the headers (A2:D6)
        data_rows = [
            [student_ws.cell(row=row, column=col).value for col in range(1, 5)]
            for row in range(2, 7)  # Rows 2 through 6 (A2:D6)
        ]

        # Ensure that each row has data (not None)
        non_empty_rows = [row for row in data_rows if all(cell is not None for cell in row)]
        if len(non_empty_rows) == expected_row_count:
            score += 2  # Award 2 points for having the correct number of data rows
        else:
            feedback.append("Incorrect number of data rows")

        # Now compare the content of the student's data with the expected solution data
        solution_data = [
            ["CustomerID", "FirstName", "LastName", "Email"],
            [101, "John", "Doe", "johndoe@example.com"],
            [102, "Jane", "Smith", "janesmith@example.com"],
            [103, "Michael", "Johnson", "mjohnson@example.com"],
            [104, "Peter", "Parker", "pparker@dailybugle.com"],
            [105, "Tony", "Stark", "tstark@starkindustries.com"]
        ]

        # Compare the content row by row (data only, ignoring the headers)
        matching_cells = 0
        total_cells = 20  # 5 rows * 4 columns

        for r in range(5):
            for c in range(4):
                solution_value = solution_data[r + 1][c]  # Offset by 1 to skip header
                student_value = data_rows[r][c]

                if student_value == solution_value:
                    matching_cells += 1
                else:
                    feedback.append("Imported data is incorrect in 1(or more) rows.")

        # Award points based on the number of correctly matched cells
        content_points = (matching_cells / total_cells) * 6  # 6 points for content accuracy
        score += content_points

        return score, total_points, feedback
    except Exception as e:
        print(f"Error comparing workbooks: {e}")
        return 0, total_points  # In case of an error, give 0 score but consider total points
    
def grade_challenge_2(student_path):
    try:
        # Load the student workbook and select the active sheet
        student_wb = openpyxl.load_workbook(student_path)
        student_ws = student_wb.active

        # Initialize scoring variables
        score = 0
        total_points = 15  # Adjust based on grading
        feedback = []

        # 1. Searching for data (4 points)
        # Check if the cell with the employee making $75,000 is highlighted in yellow
        target_cell = "B144"  # The cell containing the employee's name making $75,000
        if student_ws[target_cell].fill.start_color.rgb == "FFFFFF00":  # Yellow
            score += 2
        else:
            feedback.append("Cell B144 (name of employee with $75,000 salary) is not highlighted in yellow.")

        # 2. Navigating to named cells/ranges (5 points)
        # Check if "EmployeeInfo" named range exists
        if "EmployeeInfo" in student_wb.defined_names:
            score += 2
        else:
            feedback.append("Named range 'EmployeeInfo' not found.")

        # Check if the named range "EmployeeInfo" has the correct range A1:B201
        if "EmployeeInfo" in student_wb.defined_names:
            defined_range = student_wb.defined_names["EmployeeInfo"].attr_text
            if defined_range.endswith("!$A$1:$B$201"):
                score += 2
            else:
                feedback.append("Named range 'EmployeeInfo' does not refer to cells A1:B201.")

        # Check if the font for the range is Times New Roman
        cells_in_range = student_ws["A1:B201"]
        font_correct = all(cell.font.name == "Times New Roman" for row in cells_in_range for cell in row)
        if font_correct:
            score += 2
        else:
            feedback.append("Font for 'EmployeeInfo' named range is not set to Times New Roman.")

        # 3. Hyperlinks (6 points)
        # Check if the hyperlink in cell B204 has been removed
        if not student_ws["B204"].hyperlink:
            score += 3
        else:
            feedback.append("Hyperlink in cell B204 has not been removed.")

        # Check if the hyperlink was added to cell F4 with the correct URL and display text
        cell_f4 = student_ws["F4"]
        if cell_f4.hyperlink and cell_f4.hyperlink.target == "https://www.examplecompany.com" and cell_f4.value == "Example Company":
            score += 3
        else:
            feedback.append("Cell F4 does not have the correct hyperlink and display text.")

        return score, total_points, feedback

    except Exception as e:
        print(f"Error comparing workbooks for Assignment 2: {e}")
        traceback.print_exc()
        return 0, total_points, ["An error occurred during grading."]

def grade_challenge_3_1(student_path):
    try:
        # Load workbooks and select active sheets
        student_wb = openpyxl.load_workbook(student_path)
        student_ws = student_wb.active

        # Initialize scoring variables
        score = 0
        total_points = 20  # Adjust based on grading
        feedback = []  
        
        # 1. Page Setup (4 points)
        # Check page orientation
        if student_ws.page_setup.orientation == "landscape":
            score += 1
        else:
            feedback.append("Incorrect page orientation")
            
            
        #DEBUGGING
        print("Debug - Page Setup Scaling Settings:")
        print("Orientation:", student_ws.page_setup.orientation)
        print("Fit to Width:", student_ws.page_setup.fitToWidth)
        print("Margins - Left:", student_ws.page_margins.left, "Right:", student_ws.page_margins.right)
        print("Margins - Top:", student_ws.page_margins.top, "Bottom:", student_ws.page_margins.bottom)
        print("Row Height for Row 1:", student_ws.row_dimensions[1].height)
        print("Column Width for Column A:", student_ws.column_dimensions["A"].width)

            
        # Check fit to width/height with defaults
        fit_to_width = student_ws.page_setup.fitToWidth or 1
        fit_to_height = student_ws.page_setup.fitToHeight or 1
        if fit_to_width == 1 and fit_to_height == 1:
            score += 1
        else:
            feedback.append("Incorrect page orientation")
            
            
        # Scale to Fit to 1 page is ungradeable at this time

        # Check for narrow margins
        if (round(float(student_ws.page_margins.left), 2) == 0.25 and 
            round(float(student_ws.page_margins.right), 2) == 0.25 and
            round(float(student_ws.page_margins.top), 2) == 0.75 and 
            round(float(student_ws.page_margins.bottom), 2) == 0.75):
            score += 2
        else:
            feedback.append("Margins are not set to Narrow.")

        # 2. Row Height and Column Width (2 points)
        # Check row height
        if student_ws.row_dimensions[1].height == 30:
            score += 2
        else:
            feedback.append("Row height for the header row (Row 1) is incorrect; Expected 30 points.")
            
        # Check column widths for column A (Allows for a small tolerance to mitigate Excels float points)
        if 19 <= student_ws.column_dimensions["A"].width <=21:
            score += 1.5
        else:
            feedback.append("Incorrect column width for column A; Expected 20.")
        
        # Check column widths for column B-J (Allows for a small tolerance)    
        for col in range(2, 10):
            col_letter = get_column_letter(col)
            if not (13 <= student_ws.column_dimensions[col_letter].width <= 16.5):
                feedback.append(f"Incorrect width for column {col_letter}; Expected 15.")
                break
        else:
            score += 2

        # 3. Headers and Footers (2 points)
        # Check if there's text in the left part of the header
        if student_ws.oddHeader.left and hasattr(student_ws.oddHeader.left, 'text') and student_ws.oddHeader.left.text.strip():
            score += 2
        else:
            feedback.append("No text found in the left side of the header.")

        # Check for Date in the center part of the header
        if student_ws.oddHeader.center and hasattr(student_ws.oddHeader.center, 'text') and "&D" in student_ws.oddHeader.center.text:
            score += 1
        else:
            feedback.append("Header does not contain date in center.")

        # Check if the file name is in the right side of the header
        if student_ws.oddHeader.right and hasattr(student_ws.oddHeader.right, 'text') and "&F" in student_ws.oddHeader.right.text:
            score += 1
        else:
            feedback.append("Header does not contain file name on the right.")

        # Footer checks for page numbering
        if student_ws.oddFooter.left and hasattr(student_ws.oddFooter.left, 'text') and "&P" in student_ws.oddFooter.left.text:
            score += 1
        else:
            feedback.append("Footer does not contain page number on the left.")

        if student_ws.oddFooter.right and hasattr(student_ws.oddFooter.right, 'text') and "&N" in student_ws.oddFooter.right.text:
            score += 1
        else:
            feedback.append("Footer does not contain total number of pages on the right.")

        if student_ws.oddFooter.right and hasattr(student_ws.oddFooter.right, 'text') and "&N" in student_ws.oddFooter.right.text:
            score += 1
        else:
            feedback.append("Footer does not contain total number of pages on the right.")

        # 4. Options and Views (1 point)
        if not student_ws.sheet_view.showGridLines and not student_ws.sheet_view.showRowColHeaders:
            score += 2
        else:
            feedback.append("Gridlines or headings are not hidden.")

        return score, total_points, feedback

    except Exception as e:
        print(f"Error comparing workbooks for Assignment 3.1: {e}")
        traceback.print_exc()
        return 0, total_points, ["An error occurred during grading."]

def grade_project_1(student_path):
    #Function to help with edge cases of users sorting tables and altering data locations
    def verify_unique_countries_and_prices(analysis_sheet_values):
        expected_country_prices = {
            'Taiwan': 10.15,
            'United States': 9.24,
            'Japan': 10.75,
            'Hawaii': 18.15,
            'Hong Kong': 15.62,
            'Guatemala': 3.55,
            'China': 22.53,
            'Canada': 4.99,
            'England': 50.41,
            'Australia': 69.00,
            'Kenya': 6.91
        }
        country_prices = {}
        feedback = []
        score = 0

        for row in range(4, 15):
            country = analysis_sheet_values.cell(row=row, column=1).value
            price = analysis_sheet_values.cell(row=row, column=2).value

            if country and price is not None:
                country = str(country).strip()
                try:
                    country_prices[country] = round(float(price), 2)
                except ValueError:
                    feedback.append(f"Non-numeric price for {country}: {price}")

        missing_countries = set(expected_country_prices.keys()) - set(country_prices.keys())
        extra_countries = set(country_prices.keys()) - set(expected_country_prices.keys())

        if missing_countries:
            feedback.append(f"Missing countries: {', '.join(missing_countries)}")
        if extra_countries:
            feedback.append(f"Extra countries found: {', '.join(extra_countries)}")

        price_matches = 0
        total_countries = len(expected_country_prices)

        for country, expected_price in expected_country_prices.items():
            if country in country_prices:
                if abs(country_prices[country] - expected_price) < 0.01:
                    price_matches += 1
                else:
                    feedback.append(
                        f"Incorrect price for {country}. Expected {expected_price}, Got {country_prices[country]}"
                    )

        country_score = 5 if len(country_prices) == total_countries else 1
        price_score = (price_matches / total_countries) * 3
        score = country_score + price_score
        return score, feedback

    
    def verify_unique_countries_and_ratings(analysis_sheet):
        # Dictionary of expected countries and ratings
        expected_country_ratings = {
            'Taiwan': 93.64,
            'United States': 93.24,
            'Japan': 92.38,
            'Hawaii': 93.42,
            'Hong Kong': 92.67,
            'Guatemala': 90.5,
            'China': 90,
            'Canada': 93.6,
            'England': 94.5,
            'Australia': 96,
            'Kenya': 94
        }
        
        # Extract unique countries and their ratings from the sheet
        country_ratings = {}
        
        # Assuming the data is in columns D and E, starting from row 4 to 14
        for row in range(4, 15):
            country = analysis_sheet.cell(row=row, column=4).value
            rating = analysis_sheet.cell(row=row, column=5).value
            
            if country and rating is not None:
                # Normalize country names (strip whitespace, handle potential capitalization issues)
                country = str(country).strip()
                
                # Store the rating, allowing for small floating-point variations
                country_ratings[country] = round(float(rating), 2)
        
        # Check if all expected countries are present
        missing_countries = set(expected_country_ratings.keys()) - set(country_ratings.keys())
        extra_countries = set(country_ratings.keys()) - set(expected_country_ratings.keys())
        
        feedback = []
        score = 0
        
        if missing_countries:
            feedback.append(f"Missing countries: {', '.join(missing_countries)}")
        
        if extra_countries:
            feedback.append(f"Extra countries found: {', '.join(extra_countries)}")
        
        # Check ratings for each country
        rating_matches = 0
        total_countries = len(expected_country_ratings)
        
        for country, expected_rating in expected_country_ratings.items():
            if country in country_ratings:
                # Allow a small tolerance for floating-point comparisons
                if abs(country_ratings[country] - expected_rating) < 0.01:
                    rating_matches += 1
                else:
                    feedback.append(f"Incorrect rating for {country}. Expected {expected_rating}, Got {country_ratings[country]}")
        
        # Calculate scores
        country_score = 4 if len(country_ratings) == total_countries else 2
        rating_score = (rating_matches / total_countries) * 3
        
        score = country_score + rating_score
        
        return score, feedback

    try:
        # First pass: Check formulas (data_only=False)
        wb_formulas = openpyxl.load_workbook(student_path, data_only=False)
        
        # Second pass: Check values (data_only=True)
        wb_values = openpyxl.load_workbook(student_path, data_only=True)

        # Initialize scoring variables
        score = 0
        total_points = 60  # Base points
        feedback = []

        # Define sheet to be graded
        analysis_sheet_values = wb_values["CoffeeAnalysis"]

        # Verify Unique Countries and Prices
        unique_countries_score, unique_countries_feedback = verify_unique_countries_and_prices(analysis_sheet_values)
        score += unique_countries_score
        feedback.extend(unique_countries_feedback)
        
        # Verify Unique Countries and Ratings
        unique_ratings_score, unique_ratings_feedback = verify_unique_countries_and_ratings(analysis_sheet_values)
        score += unique_ratings_score
        feedback.extend(unique_ratings_feedback)
        
        
        # Individual Calculations Grading
        calc_checks = [
            
            # Cell B15: Overall Average USD per Unit
            {
                'cell': 'B15', 
                'expected_value': 20.12,
                'points': {
                    'value': 5,
                }
            },
            # Cell E15: Overall Average Rating
            {
                'cell': 'E15', 
                'expected_value': 93.08520928987156,
                'points': {
                    'value': 5,
                }
            },
            # Cell I4: Most Expensive Country of Origin
            {
                'cell': 'I4', 
                'expected_value': "Australia",
                'points': {
                    'value': 5,
                    'formula': 5
                }
            },
            # Cell I5: Least Expensive Country of Origin
            {
                'cell': 'I5', 
                'expected_value': "Guatemala",
                'points': {
                    'value': 5,
                    'formula': 5
                }
            },
            # Cell I7: Country with the Highest Rating
            {
                'cell': 'I7', 
                'expected_value': "Australia",
                'points': {
                    'value': 5,
                    'formula': 5
                }
            },
            # Cell I8: Country with the Lowest Rating
            {
                'cell': 'I8', 
                'expected_value': "China",
                'points': {
                    'value': 5,
                    'formula': 5
                }
            },
            # Cell I12: Average Length of Reviews
            {
                'cell': 'I12', 
                'expected_value': 269.75607779578604,
                'points': {
                    'value': 5,
                    'formula': 5
                }
            },
            # Cell I13: Longest Review Length
            {
                'cell': 'I13', 
                'expected_value': 509,
                'points': {
                    'value': 5,
                    'formula': 5
                }
            },
            # Cell I14: Shortest Review Length
            {
                'cell': 'I14', 
                'expected_value': 66,
                'points': {
                    'value': 5,
                    'formula': 5
                }
            },
            
            #--------BONUS QUESTION---------------------
            # Cell I18: Country Skewing Results
            {
                'cell': 'I18', 
                'expected_value': "Australia",
                'points': {
                    'value': 5
                }
            }
            
        ]

        # Helper function to compare values with type checking
        def compare_values(actual_value, expected_value):
            if isinstance(expected_value, (int, float)):
                try:
                    if actual_value is None:
                        return False
                    actual_float = float(actual_value)
                    return round(actual_float, 2) == round(float(expected_value), 2)
                except (TypeError, ValueError):
                    return False
            else:
                # For text comparisons, convert both to strings and compare
                return str(actual_value).strip() == str(expected_value).strip()

        # Perform checks for each calculation
        for calc in calc_checks:
            cell = calc['cell']
            try:
                cell_obj = analysis_sheet_values[cell]
                cell_value = wb_values['CoffeeAnalysis'][cell].value  # Adjusted sheet name


                #-------WORK IN PROGRESS----------------
                # Formula check: Ensure any formula exists
                #if 'points' in calc and 'formula' in calc['points']:
                #    if cell_obj.data_type == 'f':  # Cell contains a formula
                #        score += calc['points']['formula']
                #    else:
                #        feedback.append(f"Cell {cell}: No formula found (cell contains a static value)")


                # Value check with type handling
                if 'expected_value' in calc:
                    value_match = compare_values(cell_value, calc['expected_value'])
                    if value_match:
                        score += calc['points']['value']
                    else:
                        value_feedback = [
                            f"Cell {cell} Value Check:",
                            f"  - Received: {cell_value}",
                            f"  - Expected: {calc['expected_value']}"
                        ]
                        feedback.extend(value_feedback)

            except Exception as e:
                print(f"Error processing cell {cell}: {e}")
                feedback.append(f"Error processing cell {cell}: {e}")

        return score, total_points, feedback

    except Exception as e:
        print(f"Error grading Project 1: {e}")
        traceback.print_exc()
        return 0, total_points, [f"An error occurred during grading: {str(e)}"]
      
def grade_project_2(student_path):
    try:
        # First pass: Check formulas (data_only=False)
        wb_formulas = openpyxl.load_workbook(student_path, data_only=False)
        # Second pass: Check values (data_only=True)
        wb_values = openpyxl.load_workbook(student_path, data_only=True)

        # Initialize scoring variables
        score = 0
        total_points = 50
        feedback = []

        # Check if all required sheets are present
        required_sheets = ["Report", "Participants", "Times", "Names & Emails"]
        sheet_names = wb_formulas.sheetnames
        if all(sheet in sheet_names for sheet in required_sheets):
            score += 4
        else:
            feedback.append("Sheet structure incorrect: Required sheets not found or incorrectly named.")

        # Sheet 1: "Report" - Cell Values Check
        report_sheet = wb_values["Report"]
        report_values = {
            "B2": 917, "B3": 283, "B4": 332,
            "D2": 574, "D3": 689, "D4": 308,
            "F2": 801, "F3": 931, "F4": 407,
            "H2": 11, "H3": 478, "H4": 70,
            "B7": 522, "B8": 49
        }
        points_per_cell = 1
        for cell, expected_value in report_values.items():
            if report_sheet[cell].value == expected_value:
                score += points_per_cell
            else:
                feedback.append(f"Incorrect value in Report sheet at {cell}. Expected {expected_value}, found {report_sheet[cell].value}.")

        # Sheet 2: "Participants"
        participants_sheet = wb_formulas["Participants"]
        # Check for the presence of a table in Participants sheet
        if len(participants_sheet.tables) > 0:
            score += 3
        else:
            feedback.append("No table found in Participants sheet.")

        # Check the number of rows in Participants sheet
        num_rows = participants_sheet.max_row
        if num_rows == 1001:
            score += 5
        elif num_rows == 523:
            score += 5
        else:
            feedback.append(f"Incorrect number of rows in Participants sheet. Found {num_rows} rows.")

        # Header Formatting Check
        try:
            header_font = participants_sheet["A1"].font
            if header_font.bold and header_font.size == 13:  # Assuming Heading 2 style has size 13 and bold
                score += 3
            else:
                feedback.append("Incorrect header formatting in Participants sheet. Expected 'Heading 2' style (bold, size 13).")
        except AttributeError:
            feedback.append("Error checking header formatting in Participants sheet.")

        # Check if top row is frozen in Participants sheet
        if participants_sheet.freeze_panes == "A2":
            score += 3
        else:
            feedback.append("Top row is not frozen in Participants sheet.")

        # Sheet 3: "Times"
        times_sheet = wb_formulas["Times"]
        # Check for the presence of a table in Times sheet
        if len(times_sheet.tables) > 0:
            score += 3
        else:
            feedback.append("No table found in Times sheet.")

        # Check the number of rows in Times sheet
        num_rows = times_sheet.max_row
        if num_rows == 523:
            score += 3
        else:
            feedback.append(f"Incorrect number of rows in Times sheet. Found {num_rows} rows.")

        # Header Formatting Check
        try:
            header_font = times_sheet["A1"].font
            if header_font.bold and header_font.size == 13:  # Assuming Heading 2 style has size 13 and bold
                score += 3
            else:
                feedback.append("Incorrect header formatting in Times sheet. Expected 'Heading 2' style (bold, size 13).")
        except AttributeError:
            feedback.append("Error checking header formatting in Times sheet.")

        # Check if top row is frozen in Times sheet
        if times_sheet.freeze_panes == "A2":
            score += 3
        else:
            feedback.append("Top row is not frozen in Times sheet.")

        # Sheet 4: "Names & Emails"
        names_emails_sheet = wb_values["Names & Emails"]
        match_names = True
        match_emails = True

        # List of cells that should be empty in the "Names & Emails" sheet, column B
        allowed_empty_cells = [
            3, 17, 18, 33, 61, 78, 79, 80, 85, 113, 127, 128, 138, 148, 153, 159, 161,
            183, 187, 190, 191, 205, 246, 250, 252, 279, 284, 289, 302, 309, 312, 329,
            347, 361, 365, 369, 387, 394, 398, 422, 442, 458, 467, 489, 490, 493, 497,
            499, 507
        ]

        for row in range(2, 524):
            participant_name = wb_values["Participants"][f"B{row}"].value
            names_emails_name = names_emails_sheet[f"A{row}"].value

            if participant_name is not None and names_emails_name != participant_name.upper():
                match_names = False
                feedback.append(f"Incorrect name format at Names & Emails sheet cell A{row}. Expected uppercase.")
                break

            participant_email = wb_values["Participants"][f"E{row}"].value
            names_emails_email = names_emails_sheet[f"B{row}"].value

            if row in allowed_empty_cells:
                if names_emails_email is not None:
                    match_emails = False
                    feedback.append(f"Cell B{row} in Names & Emails sheet should be empty but contains data.")
                    break
            else:
                if participant_email is not None and names_emails_email != participant_email.lower():
                    match_emails = False
                    feedback.append(f"Incorrect email format at Names & Emails sheet cell B{row}. Expected lowercase.")
                    break

        if match_names:
            score += 3
        if match_emails:
            score += 3

        return score, total_points, feedback

    except Exception as e:
        print(f"Error grading Project 2: {e}")
        traceback.print_exc()
        return 0, total_points, [f"An error occurred during grading: {str(e)}"]