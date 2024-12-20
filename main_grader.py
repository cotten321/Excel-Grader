import os
import pandas as pd
import customtkinter as ctk
from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, PatternFill
import threading
import subprocess

# Import the grading algorithms from grading_algorithms.py
from grading_algorithms import *

# Set the appearance mode and color theme of tkinter window
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

# Grading functions that correlate to algorithms in the 'grading_algorithms.py' file
def get_grading_function(challenge_number):
    grading_functions = {
        "Project 1: Cafe Bloom": grade_project_1,
        "Project 2: Marathon Participants": grade_project_2,
        "Skill: Import data into workbooks": grade_challenge_1_1,
        "Skill: Navigate within workbooks": grade_challenge_2,
        "Skill: Format worksheets and workbooks": grade_challenge_3_1
    }
    return grading_functions.get(challenge_number), grading_functions

# Link the users input to a called function
def process_submissions(folder_path, challenge_number, output_path, progress_callback, completion_callback):
    grading_function, _ = get_grading_function(challenge_number)
    
    #Handles if user enters wrong function
    if not grading_function:
        completion_callback(False, "No grading function available.")
        return

    #Start empty list to place grades for each student in
    grades = []
    student_folders = [f for f in os.listdir(folder_path) if os.path.isdir(os.path.join(folder_path, f))]
    # Indicates the number of folders to iterate through
    total_students = len(student_folders)


    '''
    Append the grading result for each student to the grades list. 
    If the grading function succeeds, it records the student's folder name, score, total points, percentage, and feedback.
    If an error occurs during grading (e.g., file format issue), the grade is set to 0, and an error message is added to the feedback.
    '''
    for index, student_folder in enumerate(student_folders, 1):
        student_folder_path = os.path.join(folder_path, student_folder)
        
        for file in os.listdir(student_folder_path):
            if file.endswith(".xlsx"):
                student_file_path = os.path.join(student_folder_path, file)
                print(f"Grading {student_file_path}")

                try:
                    score, total_points, feedback = grading_function(student_file_path)
                    percentage = round((score / total_points) * 100, 2) if total_points > 0 else 0
                    
                    grades.append({
                        "Student": student_folder,
                        "Score": score,
                        "Total Points": total_points,
                        "Percentage": percentage,
                        "": "",
                        "Feedback": "; ".join(feedback)
                    })
                except Exception as e:
                    grades.append({
                        "Student": student_folder,
                        "Score": 0,
                        "Total Points": total_points,
                        "Percentage": 0,
                        "": "",
                        "Feedback": f"Error: {str(e)}" # Appends the issue that caused an error with that student
                    })
        
        # Update progress (used for progress bar)
        progress = int((index / total_students) * 100)
        progress_callback(progress)

    # Export to Excel
    df = pd.DataFrame(grades)
    output_file = os.path.join(output_path, "grades_report.xlsx")
    
    # Set workbook to grading report
    wb = Workbook()
    ws = wb.active
    ws.title = "Grading Report"
    
    # Create named styles 
    outstanding_style = NamedStyle(name='Outstanding')
    outstanding_style.fill = PatternFill(start_color='CDAEED', end_color='CDAEED', fill_type='solid')
    
    good_style = NamedStyle(name='Good')
    good_style.fill = PatternFill(start_color='82EA85', end_color='82EA85', fill_type='solid')
    
    neutral_style = NamedStyle(name='Neutral')
    neutral_style.fill = PatternFill(start_color='FFFFB3', end_color='FFFFB3', fill_type='solid')
    
    bad_style = NamedStyle(name='Bad')
    bad_style.fill = PatternFill(start_color='FF4D4D', end_color='FF4D4D', fill_type='solid')
    
    # Add styles to workbook
    if 'Outstanding' not in wb.named_styles:
        wb.add_named_style(outstanding_style)
    if 'Good' not in wb.named_styles:
        wb.add_named_style(good_style)
    if 'Neutral' not in wb.named_styles:
        wb.add_named_style(neutral_style)
    if 'Bad' not in wb.named_styles:
        wb.add_named_style(bad_style)
    
    # Append the header to the grade report, with an empty column before feedback
    ws.append(["Student", "Score", "Total Points", "Percentage", "", "Feedback"])
    
    # Add student data
    for index, row in df.iterrows():
        ws.append(row.tolist())
        
        # Apply styles based on the score to the score column (B)
        cell = ws[f"B{index + 2}"]
        
        # User score over 100 (achieved bonus points)
        if row["Percentage"] > 100:
            cell.style = "Outstanding"
        # User score over 85
        elif row["Percentage"] > 85:
            cell.style = "Good"
        # User score between 70 and 85
        elif 70 <= row["Percentage"] <= 85:
            cell.style = "Neutral"
        # User score under 70
        else:
            cell.style = "Bad"
            
    # Save the report
    wb.save(output_file)
         
    # Signal completion to user with the report path
    completion_callback(True, f"Grading complete! Report saved to: {output_file}")


'''
Main Class that holds the custom Tkinter window.

Users are able to:
1. Select a folder containing student submissions,
2. Choose a grading challenge (labled as projects or skills)
3. Select an output forlder for the grading results
4. Start the grading process, calling the above functions1
5. View the progress bar, indicating completion percentage
6. Open the grading report right from the application
'''
class ExcelGraderApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        # Configure window
        self.title("Excel Grader")
        self.geometry("600x800")
        self.configure(fg_color="#F0F0F0")  # Light gray background

        # Main container
        self.main_frame = ctk.CTkFrame(
            self, 
            corner_radius=20, 
            fg_color="white", 
            bg_color="#F0F0F0"
        )
        self.main_frame.pack(pady=20, padx=20, fill="both", expand=True)

        # Title 
        self.title_label = ctk.CTkLabel(
            self.main_frame, 
            text="Excel Grader", 
            font=("San Francisco", 32, "bold"),
            text_color="#333333"
        )
        self.title_label.pack(pady=(30, 20))
        
        # Progress Bar
        self.progress_bar = ctk.CTkProgressBar(
            self.main_frame, 
            width=500, 
            height=20,
            corner_radius=10,
            fg_color="#F0F0F0",  # Light gray bar background
            progress_color="#007AFF"  # Bright blue progress
        )
        self.progress_bar.pack(pady=(10, 20))
        self.progress_bar.set(0)  # Initial state

        # Submission Folder Section
        self.create_folder_section(
            "Student Submissions", 
            self.select_submissions_folder
        )

        # Challenge Selection Section
        self.challenge_label = ctk.CTkLabel(
            self.main_frame, 
            text="Select Challenge", 
            font=("San Francisco", 16),
            text_color="#666666"
        )
        self.challenge_label.pack(anchor="w", padx=40, pady=(20, 5))

        self.challenges = [
            "Project 1: Cafe Bloom",
            "Project 2: Marathon Participants",
            "Skill: Import data into workbooks", 
            "Skill: Navigate within workbooks", 
            "Skill: Format worksheets and workbooks"
        ]

        # Creates dropdown list for users to select challenge
        self.challenge_combobox = ctk.CTkComboBox(
            self.main_frame, 
            values=self.challenges,
            width=400,
            height=40,
            border_width=1,
            border_color="#CCCCCC",
            dropdown_hover_color="#E0E0E0",
            button_hover_color="#E0E0E0",
            font=("San Francisco", 14)
        )
        self.challenge_combobox.pack(pady=10)

        # Output Folder Section
        self.create_folder_section(
            "Output Location", 
            self.select_output_folder
        )

        # Start Grading Button
        self.start_button = ctk.CTkButton(
            self.main_frame, 
            text="Start Grading", 
            command=self.start_grading,
            width=400,
            height=50,
            corner_radius=25,
            font=("San Francisco", 16, "bold"),
            fg_color="#007AFF",  # Blue
            hover_color="#0056b3"
        )
        self.start_button.pack(pady=(30, 20))
        
        # Status Label
        self.status_label = ctk.CTkLabel(
            self.main_frame, 
            text="", 
            font=("San Francisco", 14),
            text_color="#A0A0A0"
        )
        self.status_label.pack(pady=(10, 20))

        # State variables initial state (keeps track of user selcections)
        self.submissions_folder = None
        self.output_folder = None

    def create_folder_section(self, label_text, browse_command):
        # Label
        label = ctk.CTkLabel(
            self.main_frame, 
            text=label_text, 
            font=("San Francisco", 16),
            text_color="#666666"
        )
        label.pack(anchor="w", padx=40, pady=(20, 5))

        # Container for entry and button
        container = ctk.CTkFrame(
            self.main_frame, 
            fg_color="transparent"
        )
        container.pack(pady=10)

        # Entry field
        entry = ctk.CTkEntry(
            container, 
            width=330,
            height=40,
            placeholder_text=f"Select {label_text.lower()}",
            border_width=1,
            border_color="#CCCCCC",
            font=("San Francisco", 14)
        )
        entry.pack(side="left", padx=(0, 10))

        # Browse button
        browse_btn = ctk.CTkButton(
            container, 
            text="Browse", 
            command=lambda: self.browse_folder(entry, browse_command),
            width=60,
            height=40,
            corner_radius=10,
            fg_color="#F2F2F7",  # Very light gray
            text_color="#007AFF",  # Blue
            hover_color="#E0E0E5"
        )
        browse_btn.pack(side="right")

        # Store references for later use
        if label_text == "Student Submissions":
            self.submissions_entry = entry
        else:
            self.output_entry = entry
 
    def browse_folder(self, entry_widget, selection_method):
        selection_method()
        entry_widget.configure(state="normal")
        entry_widget.delete(0, "end")
        entry_widget.insert(0, self.submissions_folder if "submissions" in entry_widget.cget("placeholder_text").lower() else self.output_folder)
        entry_widget.configure(state="disabled")

    #Updates the widget to show the users selection
    def select_submissions_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            self.submissions_folder = folder
            self.submissions_entry.configure(state="normal")
            self.submissions_entry.delete(0, "end")
            self.submissions_entry.insert(0, folder)
            self.submissions_entry.configure(state="readonly")

    # Updates the widget to show the users selection
    def select_output_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            self.output_folder = folder
            self.output_entry.configure(state="normal")
            self.output_entry.delete(0, "end")
            self.output_entry.insert(0, folder)
            self.output_entry.configure(state="readonly")

    # If all inputs are present, start the grading process
    def start_grading(self):
        if not self.submissions_folder or not self.output_folder or not self.challenge_combobox.get():
            messagebox.showwarning("Input Error", "Please select all required inputs.")
            return

        # Disable start button during grading (Needed to prevent users from double calling)
        self.start_button.configure(state="disabled")
    
        self.progress_bar.set(0)
        self.status_label.configure(text="Grading in progress...")

        # Updates the top progress bar (just aesthetic)
        def progress_update(value):
            self.progress_bar.set(value / 100)

        # Resets the GUI back to the "standard" state
        def grading_complete(success, message):
            self.start_button.configure(state="normal")
            self.progress_bar.set(1 if success else 0)
            self.status_label.configure(text=message)

            if success:
                # Extract the full path of the generated report from the message
                report_path = message.split(": ")[-1]

                # Show custom completion dialog with the path to the report
                self.show_completion_dialog(report_path)

        # Start grading in a separate thread (This prevents the GUI from freezing while grading)
        threading.Thread(
            target=process_submissions, 
            args=(
                self.submissions_folder, 
                self.challenge_combobox.get(), 
                self.output_folder,
                progress_update,
                grading_complete
            ), 
            daemon=True
        ).start()
        
    # Create a dialog for grading completion with Open report and Close buttons   
    def show_completion_dialog(self, report_path):
        
        # Creates a top-level window
        dialog = ctk.CTkToplevel(self)
        dialog.title("Grading Complete")
        dialog.geometry("350x200")
        dialog.resizable(False, False)
        dialog.grab_set() # This makes the dialogue box modal (prevents the user from taking other action on main window)

        # Success message label
        message_label = ctk.CTkLabel(
            dialog, 
            text="Grading is complete!", 
            font=("San Francisco", 18, "bold"),
            text_color="#333333"
        )
        message_label.pack(pady=(30, 20))

        # Button frame
        button_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        button_frame.pack(pady=20)

        # Open Report button
        open_button = ctk.CTkButton(
            button_frame, 
            text="Open Report", 
            command=lambda: self.open_excel_report(report_path, dialog),
            width=120,
            height=40,
            corner_radius=25,
            fg_color="#007AFF",  # Blue
            hover_color="#0056b3"   #Slightly darker blue
        )
        open_button.pack(side="left", padx=10)

        # Close button
        close_button = ctk.CTkButton(
            button_frame, 
            text="Close", 
            command=dialog.destroy,
            width=120,
            height=40,
            corner_radius=25,
            fg_color="#F2F2F7",  # Light gray
            text_color="#007AFF",   # Blue
            hover_color="#E0E0E5"   # Gray
        )
        close_button.pack(side="right", padx=10)
        
    # Open the Excel report 
    def open_excel_report(self, file_path, parent_dialog=None):
        try:
            if os.name == 'nt':  # Windows
                os.startfile(file_path)
            # If user is using MacOS or Linux, (Remove if packaging as .exe)
            elif os.name == 'posix':  # macOS and Linux
                if subprocess.sys.platform == 'darwin':  # macOS
                    subprocess.call(('open', file_path))
                else:  # Linux
                    subprocess.call(('xdg-open', file_path))
                
            # Close the parent dialog if provided
            if parent_dialog:
                parent_dialog.destroy()
        except Exception as e:
            messagebox.showerror("Error", f"Could not open the report: {str(e)}")

def main():
    app = ExcelGraderApp()
    app.mainloop()

if __name__ == "__main__":
    main()